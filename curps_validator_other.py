import pandas as pd
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Diccionario completo de entidades federativas
ENTIDADES = {
    'AG': 'Aguascalientes', 'BC': 'Baja California', 'BS': 'Baja California Sur',
    'CC': 'Campeche', 'CL': 'Coahuila', 'CM': 'Colima', 'CS': 'Chiapas',
    'CH': 'Chihuahua', 'DF': 'Ciudad de México', 'DG': 'Durango', 'GT': 'Guanajuato',
    'GR': 'Guerrero', 'HG': 'Hidalgo', 'JC': 'Jalisco', 'MC': 'México',
    'MN': 'Michoacán', 'MS': 'Morelos', 'NT': 'Nayarit', 'NL': 'Nuevo León',
    'OC': 'Oaxaca', 'PL': 'Puebla', 'QT': 'Querétaro', 'QR': 'Quintana Roo',
    'SP': 'San Luis Potosí', 'SL': 'Sinaloa', 'SR': 'Sonora', 'TC': 'Tabasco',
    'TS': 'Tamaulipas', 'TL': 'Tlaxcala', 'VZ': 'Veracruz', 'YN': 'Yucatán',
    'ZS': 'Zacatecas', 'NE': 'Nacido en el Extranjero'
}

REF_DATE = datetime(2025, 7, 10)

def validar_letras(texto, cantidad):
    return len(texto) == cantidad and texto.isalpha()

def validar_numeros(texto, cantidad):
    return len(texto) == cantidad and texto.isdigit()

def validar_fecha(fecha_str):
    if not validar_numeros(fecha_str, 6):
        return False, "Formato de fecha incorrecto"
    try:
        anio = int(fecha_str[:2])
        mes = int(fecha_str[2:4])
        dia = int(fecha_str[4:6])
        if mes < 1 or mes > 12:
            return False, f"Mes inválido: {mes}"
        if dia < 1 or dia > 31:
            return False, f"Día inválido: {dia}"
        if mes in [4, 6, 9, 11] and dia > 30:
            return False, f"Día inválido para mes {mes}: {dia}"
        if mes == 2:
            siglo = 2000 if anio <= 24 else 1900
            anio_completo = siglo + anio
            if anio_completo % 4 == 0 and (anio_completo % 100 != 0 or anio_completo % 400 == 0):
                if dia > 29:
                    return False, f"Día inválido para febrero bisiesto: {dia}"
            else:
                if dia > 28:
                    return False, f"Día inválido para febrero: {dia}"
        return True, "Fecha válida"
    except ValueError:
        return False, "Error en formato de fecha"

def validar_sexo(sexo):
    if sexo in ['H', 'M']:
        return True, 'Hombre' if sexo == 'H' else 'Mujer'
    return False, f"Sexo inválido: {sexo}"

def validar_entidad(entidad):
    if entidad in ENTIDADES:
        return True, ENTIDADES[entidad]
    return False, f"Entidad inválida: {entidad}"

def extraer_todos_datos_13(curp):
    """Extrae todos los datos de los primeros 13 caracteres sin importar si están mal"""
    if not isinstance(curp, str):
        curp = str(curp)
    curp_limpia = curp.strip().upper()
    len_curp = len(curp_limpia)
    
    # Información básica
    info = {
        'curp_original': curp,
        'curp_limpia': curp_limpia,
        'longitud_total': len_curp,
        'primeros_13_completos': curp_limpia[:13] if len_curp >= 13 else curp_limpia,
        'longitud_primeros_13': len(curp_limpia[:13]) if len_curp >= 13 else len_curp
    }
    
    # Extraer primeros 13 caracteres
    primeros_13 = curp_limpia[:13] if len_curp >= 13 else curp_limpia
    
    # Componentes de los primeros 13 (extraer todo sin importar si está mal)
    componentes_13 = {
        'primer_apellido': primeros_13[0] if len(primeros_13) > 0 else '',
        'vocal_interna': primeros_13[1] if len(primeros_13) > 1 else '',
        'segundo_apellido': primeros_13[2] if len(primeros_13) > 2 else '',
        'nombre': primeros_13[3] if len(primeros_13) > 3 else '',
        'fecha': primeros_13[4:10] if len(primeros_13) >= 10 else '',
        'sexo': primeros_13[10] if len(primeros_13) > 10 else '',
        'entidad': primeros_13[11:13] if len(primeros_13) >= 13 else ''
    }
    
    # Validar cada componente y marcar si está mal
    validaciones = {}
    
    # 1. Letras iniciales (primeros 4 caracteres)
    letras_iniciales = primeros_13[:4]
    validaciones['letras_iniciales'] = validar_letras(letras_iniciales, 4)
    
    # 2. Fecha de nacimiento (siempre extraer, validar después)
    if len(componentes_13['fecha']) == 6:
        fecha_valida, error_fecha = validar_fecha(componentes_13['fecha'])
        validaciones['fecha'] = fecha_valida
        # Intentar calcular edad incluso si la fecha está mal
        try:
            anio = int(componentes_13['fecha'][:2])
            siglo = 2000 if anio <= 24 else 1900
            fecha_nac = datetime(
                siglo + anio,
                int(componentes_13['fecha'][2:4]),
                int(componentes_13['fecha'][4:6])
            )
            info['fecha_formateada'] = fecha_nac.strftime('%d/%m/%Y')
            info['edad'] = REF_DATE.year - fecha_nac.year
            if (REF_DATE.month, REF_DATE.day) < (fecha_nac.month, fecha_nac.day):
                info['edad'] -= 1
        except:
            info['fecha_formateada'] = 'Fecha inválida'
            info['edad'] = 'N/A'
    else:
        validaciones['fecha'] = False
        info['fecha_formateada'] = 'Fecha incompleta'
        info['edad'] = 'N/A'
    
    # 3. Sexo
    if len(componentes_13['sexo']) == 1:
        sexo_valido, sexo_nombre = validar_sexo(componentes_13['sexo'])
        validaciones['sexo'] = sexo_valido
        info['sexo_legible'] = sexo_nombre
    else:
        validaciones['sexo'] = False
        info['sexo_legible'] = 'Sexo faltante'
    
    # 4. Entidad federativa
    if len(componentes_13['entidad']) == 2:
        entidad_valida, entidad_nombre = validar_entidad(componentes_13['entidad'])
        validaciones['entidad'] = entidad_valida
        info['entidad_legible'] = entidad_nombre
    else:
        validaciones['entidad'] = False
        info['entidad_legible'] = 'Entidad incompleta'
    
    # Agregar todos los componentes al resultado (sin importar si están mal)
    info.update({
        'primer_apellido': componentes_13['primer_apellido'],
        'vocal_interna': componentes_13['vocal_interna'],
        'segundo_apellido': componentes_13['segundo_apellido'],
        'nombre': componentes_13['nombre'],
        'fecha_nacimiento': componentes_13['fecha'],
        'sexo_codigo': componentes_13['sexo'],
        'entidad_codigo': componentes_13['entidad'],
        'consonantes': curp_limpia[13:16] if len_curp >= 16 else '',
        'homoclave': curp_limpia[16:18] if len_curp >= 18 else '',
        'validaciones': validaciones
    })
    
    return info

def crear_reporte_completo(resultados):
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Datos Primeros 13"
    ws_resumen = wb.create_sheet("Resumen", 0)
    
    # Estilos
    fill_valido = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_invalido = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    fill_normal = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    font_valido = Font(color='006100', bold=True)
    font_invalido = Font(color='9C0006', bold=True)
    font_normal = Font(color='000000')
    
    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Hoja de resumen
    if ws_resumen is not None:
        ws_resumen['A1'] = 'EXTRACCIÓN COMPLETA DE DATOS CURP'
        ws_resumen['A1'].font = Font(bold=True, size=16)
        ws_resumen['A3'] = 'INFORMACIÓN EXTRAÍDA'
        ws_resumen['A3'].font = Font(bold=True, size=12)
        ws_resumen['A5'] = 'Total de registros:'
        ws_resumen['B5'] = len(resultados)
        ws_resumen['A6'] = 'CURPs con 13+ caracteres:'
        ws_resumen['B6'] = sum(1 for r in resultados if r['longitud_primeros_13'] >= 13)
        ws_resumen['A7'] = 'CURPs con menos de 13 caracteres:'
        ws_resumen['B7'] = sum(1 for r in resultados if r['longitud_primeros_13'] < 13)
        
        ws_resumen['A9'] = 'NOTA:'
        ws_resumen['A9'].font = Font(bold=True, size=12)
        ws_resumen['A10'] = 'Este reporte extrae TODOS los datos de los primeros 13 caracteres'
        ws_resumen['A11'] = 'Los datos en ROJO están marcados como inválidos'
        ws_resumen['A12'] = 'Los datos en VERDE están marcados como válidos'
        ws_resumen['A13'] = 'Se extrae información incluso si las fechas están mal'
    
    # Hoja de datos principales
    columnas = [
        'CURP Original', 'Longitud Total', 'Primeros 13 Completos', 'Longitud Primeros 13',
        '1er Apellido', 'Vocal', '2do Apellido', 'Nombre',
        'Fecha Nac', 'Fecha Formato', 'Edad', 'Sexo Código', 'Sexo', 
        'Entidad Código', 'Entidad', 'Consonantes', 'Homoclave'
    ]
    
    if ws is not None:
        for col_num, col_name in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_num, result in enumerate(resultados, 2):
            validaciones = result.get('validaciones', {})
            
            for col_num, value in enumerate([
                result['curp_original'],
                result['longitud_total'],
                result['primeros_13_completos'],
                result['longitud_primeros_13'],
                result['primer_apellido'],
                result['vocal_interna'],
                result['segundo_apellido'],
                result['nombre'],
                result['fecha_nacimiento'],
                result.get('fecha_formateada', ''),
                result.get('edad', ''),
                result['sexo_codigo'],
                result.get('sexo_legible', ''),
                result['entidad_codigo'],
                result.get('entidad_legible', ''),
                result['consonantes'],
                result['homoclave']
            ], 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Aplicar colores según validación
                if col_num == 5:  # 1er Apellido
                    if validaciones.get('letras_iniciales', False):
                        cell.fill = fill_valido
                        cell.font = font_valido
                    else:
                        cell.fill = fill_invalido
                        cell.font = font_invalido
                elif col_num == 6:  # Vocal
                    if validaciones.get('letras_iniciales', False):
                        cell.fill = fill_valido
                        cell.font = font_valido
                    else:
                        cell.fill = fill_invalido
                        cell.font = font_invalido
                elif col_num == 7:  # 2do Apellido
                    if validaciones.get('letras_iniciales', False):
                        cell.fill = fill_valido
                        cell.font = font_valido
                    else:
                        cell.fill = fill_invalido
                        cell.font = font_invalido
                elif col_num == 8:  # Nombre
                    if validaciones.get('letras_iniciales', False):
                        cell.fill = fill_valido
                        cell.font = font_valido
                    else:
                        cell.fill = fill_invalido
                        cell.font = font_invalido
                elif col_num == 9:  # Fecha Nac
                    if validaciones.get('fecha', False):
                        cell.fill = fill_valido
                        cell.font = font_valido
                    else:
                        cell.fill = fill_invalido
                        cell.font = font_invalido
                elif col_num == 10:  # Fecha Formato
                    if validaciones.get('fecha', False):
                        cell.fill = fill_valido
                        cell.font = font_valido
                    else:
                        cell.fill = fill_invalido
                        cell.font = font_invalido
                elif col_num == 11:  # Edad
                    if validaciones.get('fecha', False):
                        cell.fill = fill_valido
                        cell.font = font_valido
                    else:
                        cell.fill = fill_invalido
                        cell.font = font_invalido
                elif col_num == 12:  # Sexo Código
                    if validaciones.get('sexo', False):
                        cell.fill = fill_valido
                        cell.font = font_valido
                    else:
                        cell.fill = fill_invalido
                        cell.font = font_invalido
                elif col_num == 13:  # Sexo
                    if validaciones.get('sexo', False):
                        cell.fill = fill_valido
                        cell.font = font_valido
                    else:
                        cell.fill = fill_invalido
                        cell.font = font_invalido
                elif col_num == 14:  # Entidad Código
                    if validaciones.get('entidad', False):
                        cell.fill = fill_valido
                        cell.font = font_valido
                    else:
                        cell.fill = fill_invalido
                        cell.font = font_invalido
                elif col_num == 15:  # Entidad
                    if validaciones.get('entidad', False):
                        cell.fill = fill_valido
                        cell.font = font_valido
                    else:
                        cell.fill = fill_invalido
                        cell.font = font_invalido
                else:  # Otros campos (CURP Original, Longitud, etc.)
                    cell.fill = fill_normal
                    cell.font = font_normal
        
        # Ajustar ancho de columnas
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in col:
                if cell.value is not None:
                    try:
                        value_str = str(cell.value)
                        value_length = len(value_str)
                        if value_length > max_length:
                            max_length = value_length
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Filtros automáticos
        if ws.dimensions:
            ws.auto_filter.ref = ws.dimensions
    
    return wb

# Leer el archivo Excel
try:
    df = pd.read_excel('Aspirantes.xlsx', sheet_name='Aspirantes')
    print(f"Archivo leído correctamente. Total de registros: {len(df)}")
except Exception as e:
    print(f"Error al leer el archivo: {e}")
    exit(1)

print("Extrayendo TODOS los datos de los primeros 13 caracteres...")
resultados = []
for index, row in df.iterrows():
    curp = row['CURP']
    datos = extraer_todos_datos_13(curp)
    resultados.append(datos)

print(f"\n=== EXTRACCIÓN COMPLETA DE DATOS ===")
print(f"Total de registros procesados: {len(resultados)}")
print(f"CURPs con 13+ caracteres: {sum(1 for r in resultados if r['longitud_primeros_13'] >= 13)}")
print(f"CURPs con menos de 13 caracteres: {sum(1 for r in resultados if r['longitud_primeros_13'] < 13)}")

# Crear y guardar reporte completo
wb = crear_reporte_completo(resultados)
nombre_archivo = 'Extraccion_Completa_CURP.xlsx'
wb.save(nombre_archivo)
print(f"\nReporte generado: '{nombre_archivo}'")
print("El archivo incluye:")
print("- TODOS los datos de los primeros 13 caracteres")
print("- Información extraída sin importar si está mal")
print("- Datos válidos en VERDE")
print("- Datos inválidos en ROJO")
print("- Filtros para análisis rápido")
print("- Se extrae información incluso con fechas mal") 