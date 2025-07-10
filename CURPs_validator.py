import pandas as pd
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Diccionario completo de entidades federativas
ENTIDADES = {
    'AG': 'Aguascalientes', 'BC': 'Baja California', 'BS': 'Baja California Sur',
    'CC': 'Campeche', 'CL': 'Coahuila', 'CM': 'Colima', 'CS': 'Chiapas',
    'CH': 'Chihuahua', 'DF': 'Ciudad de México', 'DG': 'Durango', 'GT': 'Guanajuato',
    'GR': 'Guerrero', 'HG': 'Hidalgo', 'JC': 'Jalisco', 'MC': 'México',
    'MN': 'Michoacán', 'MS': 'Morelos', 'NT': 'Nayarit', 'NL': 'Nuevo León',
    'OC': 'Oaxaca', 'PL': 'Puebla', 'QT': 'Querétaro', 'QR': 'Quintana Roo',
    'SP': 'San Luis Potosí', 'SL': 'Sinaloa', 'SR': 'Sonora', 'TC': 'Tabasco',
    'TS': 'Tamaulipas', 'TL': 'Tlaxcala', 'VZ': 'Veracruz', 'YN': 'Yucatán',
    'ZS': 'Zacatecas', 'NE': 'Nacido en el Extranjero'
}

REF_DATE = datetime(2025, 7, 10)

def validar_letras(texto, cantidad):
    return len(texto) == cantidad and texto.isalpha()

def validar_numeros(texto, cantidad):
    return len(texto) == cantidad and texto.isdigit()

def validar_fecha(fecha_str):
    if not validar_numeros(fecha_str, 6):
        return False, "Formato de fecha incorrecto"
    try:
        anio = int(fecha_str[:2])
        mes = int(fecha_str[2:4])
        dia = int(fecha_str[4:6])
        if mes < 1 or mes > 12:
            return False, f"Mes inválido: {mes}"
        if dia < 1 or dia > 31:
            return False, f"Día inválido: {dia}"
        if mes in [4, 6, 9, 11] and dia > 30:
            return False, f"Día inválido para mes {mes}: {dia}"
        if mes == 2:
            siglo = 2000 if anio <= 24 else 1900
            anio_completo = siglo + anio
            if anio_completo % 4 == 0 and (anio_completo % 100 != 0 or anio_completo % 400 == 0):
                if dia > 29:
                    return False, f"Día inválido para febrero bisiesto: {dia}"
            else:
                if dia > 28:
                    return False, f"Día inválido para febrero: {dia}"
        return True, "Fecha válida"
    except ValueError:
        return False, "Error en formato de fecha"

def validar_sexo(sexo):
    if sexo in ['H', 'M']:
        return True, 'Hombre' if sexo == 'H' else 'Mujer'
    return False, f"Sexo inválido: {sexo}"

def validar_entidad(entidad):
    if entidad in ENTIDADES:
        return True, ENTIDADES[entidad]
    return False, f"Entidad inválida: {entidad}"

def analizar_primeros_13(curp):
    """Analiza específicamente los primeros 13 caracteres de la CURP"""
    if not isinstance(curp, str):
        curp = str(curp)
    curp_limpia = curp.strip().upper()
    len_curp = len(curp_limpia)
    
    # Información básica
    info = {
        'curp_original': curp,
        'curp_limpia': curp_limpia,
        'longitud_total': len_curp,
        'estado_completo': 'INVÁLIDA',
        'estado_primeros_13': 'INVÁLIDOS',
        'componentes_validos_13': [],
        'errores_primeros_13': []
    }
    
    # Extraer primeros 13 caracteres
    primeros_13 = curp_limpia[:13] if len_curp >= 13 else curp_limpia
    info['longitud_primeros_13'] = len(primeros_13)
    
    # Componentes de los primeros 13
    componentes_13 = {
        'primer_apellido': primeros_13[0] if len(primeros_13) > 0 else '',
        'vocal_interna': primeros_13[1] if len(primeros_13) > 1 else '',
        'segundo_apellido': primeros_13[2] if len(primeros_13) > 2 else '',
        'nombre': primeros_13[3] if len(primeros_13) > 3 else '',
        'fecha': primeros_13[4:10] if len(primeros_13) >= 10 else '',
        'sexo': primeros_13[10] if len(primeros_13) > 10 else '',
        'entidad': primeros_13[11:13] if len(primeros_13) >= 13 else ''
    }
    
    # Validar cada componente de los primeros 13
    validaciones_13 = {}
    
    # 1. Letras iniciales (primeros 4 caracteres)
    letras_iniciales = primeros_13[:4]
    if validar_letras(letras_iniciales, 4):
        validaciones_13['letras_iniciales'] = True
        info['componentes_validos_13'].append('Letras iniciales')
    else:
        validaciones_13['letras_iniciales'] = False
        info['errores_primeros_13'].append('Error en letras iniciales')
    
    # 2. Fecha de nacimiento
    if len(componentes_13['fecha']) == 6:
        fecha_valida, error_fecha = validar_fecha(componentes_13['fecha'])
        validaciones_13['fecha'] = fecha_valida
        if fecha_valida:
            info['componentes_validos_13'].append('Fecha de nacimiento')
            # Calcular edad
            try:
                anio = int(componentes_13['fecha'][:2])
                siglo = 2000 if anio <= 24 else 1900
                fecha_nac = datetime(
                    siglo + anio,
                    int(componentes_13['fecha'][2:4]),
                    int(componentes_13['fecha'][4:6])
                )
                info['fecha_formateada'] = fecha_nac.strftime('%d/%m/%Y')
                info['edad'] = REF_DATE.year - fecha_nac.year
                if (REF_DATE.month, REF_DATE.day) < (fecha_nac.month, fecha_nac.day):
                    info['edad'] -= 1
            except:
                pass
        else:
            info['errores_primeros_13'].append(f'Fecha inválida: {error_fecha}')
    else:
        validaciones_13['fecha'] = False
        info['errores_primeros_13'].append('Fecha incompleta o faltante')
    
    # 3. Sexo
    if len(componentes_13['sexo']) == 1:
        sexo_valido, sexo_nombre = validar_sexo(componentes_13['sexo'])
        validaciones_13['sexo'] = sexo_valido
        if sexo_valido:
            info['componentes_validos_13'].append('Sexo')
            info['sexo_legible'] = sexo_nombre
        else:
            info['errores_primeros_13'].append(f'Sexo inválido: {componentes_13["sexo"]}')
    else:
        validaciones_13['sexo'] = False
        info['errores_primeros_13'].append('Sexo faltante')
    
    # 4. Entidad federativa
    if len(componentes_13['entidad']) == 2:
        entidad_valida, entidad_nombre = validar_entidad(componentes_13['entidad'])
        validaciones_13['entidad'] = entidad_valida
        if entidad_valida:
            info['componentes_validos_13'].append('Entidad federativa')
            info['entidad_legible'] = entidad_nombre
        else:
            info['errores_primeros_13'].append(f'Entidad inválida: {componentes_13["entidad"]}')
    else:
        validaciones_13['entidad'] = False
        info['errores_primeros_13'].append('Entidad incompleta o faltante')
    
    # Determinar estado de los primeros 13
    componentes_validos = len(info['componentes_validos_13'])
    if componentes_validos >= 4:  # Al menos letras, fecha, sexo y entidad
        info['estado_primeros_13'] = 'VÁLIDOS'
    elif componentes_validos >= 2:
        info['estado_primeros_13'] = 'PARCIALMENTE VÁLIDOS'
    else:
        info['estado_primeros_13'] = 'INVÁLIDOS'
    
    # Validar CURP completa (18 caracteres)
    if len_curp == 18:
        # Validar formato completo: 4 letras + 6 números + H/M + 5 letras + 2 alfanuméricos
        if re.match(r'^[A-Z]{4}[0-9]{6}[HM][A-Z]{5}[0-9A-Z]{2}$', curp_limpia):
            # Validar que los primeros 13 sean correctos
            if all([validaciones_13['letras_iniciales'], validaciones_13['fecha'], 
                   validaciones_13['sexo'], validaciones_13['entidad']]):
                info['estado_completo'] = 'VÁLIDA'
            else:
                info['estado_completo'] = 'INVÁLIDA (error en primeros 13)'
        else:
            info['estado_completo'] = 'INVÁLIDA (formato incorrecto)'
    else:
        info['estado_completo'] = f'INVÁLIDA (longitud: {len_curp})'
    
    # Agregar componentes individuales al resultado
    info.update({
        'primer_apellido': componentes_13['primer_apellido'],
        'vocal_interna': componentes_13['vocal_interna'],
        'segundo_apellido': componentes_13['segundo_apellido'],
        'nombre': componentes_13['nombre'],
        'fecha_nacimiento': componentes_13['fecha'],
        'sexo_codigo': componentes_13['sexo'],
        'entidad_codigo': componentes_13['entidad'],
        'consonantes': curp_limpia[13:16] if len_curp >= 16 else '',
        'homoclave': curp_limpia[16:18] if len_curp >= 18 else ''
    })
    
    return info

def crear_reporte_detallado(resultados, resumen):
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Análisis Primeros 13"
    ws_resumen = wb.create_sheet("Resumen", 0)
    
    # Estilos
    fill_valido = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_parcial = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    fill_invalido = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    font_valido = Font(color='006100', bold=True)
    font_parcial = Font(color='9C5700', bold=True)
    font_invalido = Font(color='9C0006', bold=True)
    font_normal = Font(color='000000')
    
    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Hoja de resumen
    if ws_resumen is not None:
        ws_resumen['A1'] = 'RESUMEN DE ANÁLISIS DE CURP'
        ws_resumen['A1'].font = Font(bold=True, size=16)
        ws_resumen['A3'] = 'ESTADÍSTICAS GENERALES'
        ws_resumen['A3'].font = Font(bold=True, size=12)
        ws_resumen['A5'] = 'Total de registros:'
        ws_resumen['B5'] = resumen['total']
        ws_resumen['A6'] = 'CURPs válidas (18 caracteres):'
        ws_resumen['B6'] = resumen['validas_18']
        ws_resumen['A7'] = 'Primeros 13 válidos:'
        ws_resumen['B7'] = resumen['primeros_13_validos']
        ws_resumen['A8'] = 'Primeros 13 parcialmente válidos:'
        ws_resumen['B8'] = resumen['primeros_13_parciales']
        ws_resumen['A9'] = 'Primeros 13 inválidos:'
        ws_resumen['B9'] = resumen['primeros_13_invalidos']
        
        ws_resumen['A11'] = 'PORCENTAJES'
        ws_resumen['A11'].font = Font(bold=True, size=12)
        ws_resumen['A13'] = 'CURPs válidas:'
        ws_resumen['B13'] = f"{resumen['validas_18']/resumen['total']*100:.1f}%"
        ws_resumen['A14'] = 'Primeros 13 válidos:'
        ws_resumen['B14'] = f"{resumen['primeros_13_validos']/resumen['total']*100:.1f}%"
        ws_resumen['A15'] = 'Primeros 13 parciales:'
        ws_resumen['B15'] = f"{resumen['primeros_13_parciales']/resumen['total']*100:.1f}%"
        ws_resumen['A16'] = 'Primeros 13 inválidos:'
        ws_resumen['B16'] = f"{resumen['primeros_13_invalidos']/resumen['total']*100:.1f}%"
    
    # Hoja de análisis detallado
    columnas = [
        'CURP Original', 'Longitud Total', 'Estado Completo', 'Estado Primeros 13',
        '1er Apellido', 'Vocal', '2do Apellido', 'Nombre',
        'Fecha Nac', 'Fecha Formato', 'Edad', 'Sexo Código', 'Sexo', 
        'Entidad Código', 'Entidad', 'Consonantes', 'Homoclave',
        'Componentes Válidos (13)', 'Errores Primeros 13'
    ]
    
    if ws is not None:
        for col_num, col_name in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_num, result in enumerate(resultados, 2):
            # Determinar estilo basado en estado de primeros 13
            if result['estado_primeros_13'] == 'VÁLIDOS':
                fill = fill_valido
                font = font_valido
            elif result['estado_primeros_13'] == 'PARCIALMENTE VÁLIDOS':
                fill = fill_parcial
                font = font_parcial
            else:
                fill = fill_invalido
                font = font_invalido
            
            # Preparar datos
            componentes_validos = ', '.join(result.get('componentes_validos_13', []))
            errores = ', '.join(result.get('errores_primeros_13', []))
            
            datos = [
                result['curp_original'],
                result['longitud_total'],
                result['estado_completo'],
                result['estado_primeros_13'],
                result['primer_apellido'],
                result['vocal_interna'],
                result['segundo_apellido'],
                result['nombre'],
                result['fecha_nacimiento'],
                result.get('fecha_formateada', ''),
                result.get('edad', ''),
                result['sexo_codigo'],
                result.get('sexo_legible', ''),
                result['entidad_codigo'],
                result.get('entidad_legible', ''),
                result['consonantes'],
                result['homoclave'],
                componentes_validos,
                errores
            ]
            
            for col_num, value in enumerate(datos, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = fill
                cell.font = font if col_num <= 4 else font_normal
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Ajustar ancho de columnas
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in col:
                if cell.value is not None:
                    try:
                        value_str = str(cell.value)
                        value_length = len(value_str)
                        if value_length > max_length:
                            max_length = value_length
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Filtros automáticos
        if ws.dimensions:
            ws.auto_filter.ref = ws.dimensions
    
    return wb

# Leer el archivo Excel
try:
    df = pd.read_excel('Aspirantes.xlsx', sheet_name='Aspirantes')
    print(f"Archivo leído correctamente. Total de registros: {len(df)}")
except Exception as e:
    print(f"Error al leer el archivo: {e}")
    exit(1)

print("Procesando análisis de primeros 13 caracteres...")
resultados = []
for index, row in df.iterrows():
    curp = row['CURP']
    analisis = analizar_primeros_13(curp)
    resultados.append(analisis)

# Estadísticas detalladas
resumen = {
    'total': len(resultados),
    'validas_18': sum(1 for r in resultados if r['estado_completo'] == 'VÁLIDA'),
    'primeros_13_validos': sum(1 for r in resultados if r['estado_primeros_13'] == 'VÁLIDOS'),
    'primeros_13_parciales': sum(1 for r in resultados if r['estado_primeros_13'] == 'PARCIALMENTE VÁLIDOS'),
    'primeros_13_invalidos': sum(1 for r in resultados if r['estado_primeros_13'] == 'INVÁLIDOS'),
}

print(f"\n=== ANÁLISIS PRIMEROS 13 CARACTERES ===")
print(f"Total de registros: {resumen['total']}")
print(f"CURPs válidas (18 caracteres): {resumen['validas_18']} ({resumen['validas_18']/resumen['total']*100:.1f}%)")
print(f"Primeros 13 válidos: {resumen['primeros_13_validos']} ({resumen['primeros_13_validos']/resumen['total']*100:.1f}%)")
print(f"Primeros 13 parcialmente válidos: {resumen['primeros_13_parciales']} ({resumen['primeros_13_parciales']/resumen['total']*100:.1f}%)")
print(f"Primeros 13 inválidos: {resumen['primeros_13_invalidos']} ({resumen['primeros_13_invalidos']/resumen['total']*100:.1f}%)")

# Crear y guardar reporte detallado
wb = crear_reporte_detallado(resultados, resumen)
nombre_archivo = 'Analisis_Primeros_13_CURP.xlsx'
wb.save(nombre_archivo)
print(f"\nReporte generado: '{nombre_archivo}'")
print("El archivo incluye:")
print("- Todos los datos de los primeros 13 caracteres")
print("- Estado de validación completa (18 caracteres)")
print("- Componentes individuales de cada CURP")
print("- Información clara y fácil de leer")
print("- Filtros para análisis rápido")